from __future__ import annotations

import io
from typing import Optional

import structlog
from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.api.dependencies.db import get_db
from backend.auth.firebase_auth import verify_firebase_token, FirebaseUser
from backend.api.schemas.bulk_upload import UploadResponse

log = structlog.get_logger(__name__)
router = APIRouter(prefix="/api/bulk-uploads", tags=["bulk-uploads"])


# ---------------------------------------------------------------------------
# Existing endpoints (preserved exactly)
# ---------------------------------------------------------------------------

@router.post("/scalar-results", response_model=UploadResponse)
async def upload_scalar_results(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload a Solution Chemistry Excel file and upsert scalar results."""
    import sys  # noqa: PLC0415
    from types import ModuleType  # noqa: PLC0415
    if "frontend.config.variable_config" not in sys.modules:
        _stub = ModuleType("frontend.config.variable_config")
        sys.modules["frontend"] = sys.modules.get("frontend", ModuleType("frontend"))
        sys.modules["frontend.config"] = sys.modules.get("frontend.config", ModuleType("frontend.config"))
        sys.modules["frontend.config.variable_config"] = _stub
    _vc = sys.modules["frontend.config.variable_config"]
    if not hasattr(_vc, "SCALAR_RESULTS_TEMPLATE_HEADERS"):
        _vc.SCALAR_RESULTS_TEMPLATE_HEADERS = {
            "measurement_date": "Date",
            "experiment_id": "Experiment ID",
            "time_post_reaction": "Time (days)",
            "description": "Description",
            "gross_ammonium_concentration_mM": "Gross Ammonium (mM)",
            "sampling_volume_mL": "Sampling Vol (mL)",
            "background_ammonium_concentration_mM": "Bkg Ammonium (mM)",
            "background_experiment_id": "Bkg Exp ID",
            "h2_concentration": "H2 Conc (ppm)",
            "gas_sampling_volume_ml": "Gas Sample Vol (mL)",
            "gas_sampling_pressure_MPa": "Gas Pressure (MPa)",
            "final_ph": "Final pH",
            "ferrous_iron_yield": "Fe2+ Yield (%)",
            "final_nitrate_concentration_mM": "Final Nitrate (mM)",
            "final_dissolved_oxygen_mg_L": "Final DO (mg/L)",
            "co2_partial_pressure_MPa": "CO2 Pressure (MPa)",
            "final_conductivity_mS_cm": "Conductivity (mS/cm)",
            "final_alkalinity_mg_L": "Alkalinity (mg/L)",
            "overwrite": "Overwrite",
        }
    from backend.services.bulk_uploads.scalar_results import ScalarResultsUploadService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors, feedbacks = ScalarResultsUploadService.bulk_upsert_from_excel_ex(
            db, file_bytes
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("scalar_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    log.info("scalar_upload", created=created, updated=updated, user=current_user.email)
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        feedbacks=feedbacks,
        message=f"Processed: {created} created, {updated} updated, {skipped} skipped",
    )


@router.post("/new-experiments", response_model=UploadResponse)
async def upload_new_experiments(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload a New Experiments Excel file."""
    from backend.services.bulk_uploads.new_experiments import NewExperimentsUploadService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors, warnings, _info = (
            NewExperimentsUploadService.bulk_upsert_from_excel(db, file_bytes)
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("new_experiments_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(created=created, updated=updated, skipped=skipped, errors=errors,
                          warnings=warnings,
                          message=f"{created} created, {updated} updated, {skipped} skipped")


@router.post("/pxrf", response_model=UploadResponse)
async def upload_pxrf(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload a pXRF CSV/Excel file and re-evaluate characterized status for affected samples."""
    import sys  # noqa: PLC0415
    from types import ModuleType  # noqa: PLC0415
    if "frontend.config.variable_config" not in sys.modules:
        _stub = ModuleType("frontend.config.variable_config")
        sys.modules["frontend"] = sys.modules.get("frontend", ModuleType("frontend"))
        sys.modules["frontend.config"] = sys.modules.get("frontend.config", ModuleType("frontend.config"))
        sys.modules["frontend.config.variable_config"] = _stub
    _vc = sys.modules["frontend.config.variable_config"]
    if not hasattr(_vc, "PXRF_REQUIRED_COLUMNS"):
        _vc.PXRF_REQUIRED_COLUMNS = {"Reading No", "Fe", "Mg", "Si", "Ni", "Cu", "Mo", "Co", "Al", "Ca", "K", "Au"}
    from backend.services.bulk_uploads.pxrf_data import PXRFUploadService  # noqa: PLC0415
    file_bytes = await file.read()

    # Lightweight extraction of reading_no values for post-upload reverse-match.
    # Mirrors the normalization in PXRFUploadService._clean_dataframe.
    # Uses openpyxl directly to avoid pandas/numpy version-mismatch issues.
    _imported_reading_nos: set[str] = set()
    try:
        from backend.services.samples import normalize_pxrf_reading_no as _norm  # noqa: PLC0415
        try:
            import openpyxl as _openpyxl  # noqa: PLC0415
            _wb = _openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            try:
                _ws = _wb.active
                _header_row = next(_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if _header_row is not None:
                    _rn_col = next(
                        (i for i, h in enumerate(_header_row) if str(h).strip() == "Reading No"), None
                    )
                    if _rn_col is not None:
                        for _row in _ws.iter_rows(min_row=2, values_only=True):
                            _v = _row[_rn_col] if _rn_col < len(_row) else None
                            if _v is not None:
                                _s = str(_v).strip()
                                if _s:
                                    _imported_reading_nos.add(_norm(_s))
            finally:
                _wb.close()
        except Exception:
            # openpyxl failed (CSV file) — fall back to pandas
            import pandas as _pd  # noqa: PLC0415
            _df_rn = _pd.read_csv(io.BytesIO(file_bytes), usecols=["Reading No"], dtype=str)
            for _v in _df_rn["Reading No"].dropna():
                _s = str(_v).strip()
                if _s:
                    _imported_reading_nos.add(_norm(_s))
    except Exception:
        pass

    try:
        created, updated, skipped, errors, svc_warnings = PXRFUploadService.ingest_from_bytes(db, file_bytes)

        # Reverse-match: re-evaluate characterized for samples whose EA pxrf_reading_no
        # overlaps with the just-ingested readings.
        reevaluated_count = 0
        if _imported_reading_nos:
            from sqlalchemy import or_ as _or  # noqa: PLC0415
            from database.models.analysis import ExternalAnalysis as _EA  # noqa: PLC0415
            from database.models.samples import SampleInfo as _SI  # noqa: PLC0415
            from backend.services.samples import (  # noqa: PLC0415
                evaluate_characterized as _eval_char,
                log_sample_modification as _log_mod,
            )

            # Build LIKE conditions matching comma-separated pxrf_reading_no field.
            # Pattern mirrors v_pxrf_characterization view in database/event_listeners.py.
            _like_conds = []
            for _rno in _imported_reading_nos:
                _like_conds.append(
                    _or(
                        _EA.pxrf_reading_no == _rno,
                        _EA.pxrf_reading_no.like(_rno + ",%"),
                        _EA.pxrf_reading_no.like("%," + _rno + ",%"),
                        _EA.pxrf_reading_no.like("%," + _rno),
                    )
                )

            affected_eas = db.query(_EA).filter(
                _EA.analysis_type == "pXRF",  # AnalysisType.PXRF.value
                _EA.sample_id.isnot(None),
                _or(*_like_conds),
            ).all()

            for _sid in {ea.sample_id for ea in affected_eas if ea.sample_id}:
                _sample = db.query(_SI).filter(_SI.sample_id == _sid).first()
                if _sample is None:
                    continue
                _old = _sample.characterized
                _new = _eval_char(db, _sid)
                if _old != _new:
                    _sample.characterized = _new
                    _log_mod(
                        db,
                        sample_id=_sid,
                        modified_by=current_user.email,
                        modification_type="update",
                        modified_table="sample_info",
                        old_values={"characterized": _old},
                        new_values={
                            "characterized": _new,
                            "reason": "Triggered by pXRF bulk upload reverse-match",
                        },
                    )
                    reevaluated_count += 1

        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("pxrf_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")

    base_msg = f"pXRF: {created} created, {updated} updated"
    message = (
        f"{base_msg}. Updated characterized status for {reevaluated_count} sample{'s' if reevaluated_count != 1 else ''}."
        if reevaluated_count > 0
        else base_msg
    )
    log.info("pxrf_upload", created=created, updated=updated, updated_characterized=reevaluated_count, user=current_user.email)
    return UploadResponse(created=created, updated=updated, skipped=skipped, errors=errors,
                          message=message, warnings=svc_warnings)


@router.post("/aeris-xrd", response_model=UploadResponse)
async def upload_aeris_xrd(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload an Aeris XRD file (time-series mineral phases)."""
    from backend.services.bulk_uploads.aeris_xrd import AerisXRDUploadService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors = AerisXRDUploadService.bulk_upsert_from_excel(db, file_bytes)
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("aeris_xrd_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(created=created, updated=updated, skipped=skipped, errors=errors,
                          message=f"XRD: {created} created, {updated} updated")


# ---------------------------------------------------------------------------
# New endpoints (Chunk C)
# ---------------------------------------------------------------------------

@router.post("/master-results", response_model=UploadResponse)
async def upload_master_results(
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """
    Master Results upload.

    - No file: sync from the configured SharePoint path (MASTER_RESULTS_PATH).
    - With file: parse the uploaded file (manual override).
    """
    from backend.services.bulk_uploads.master_bulk_upload import MasterBulkUploadService  # noqa: PLC0415
    try:
        if file is None:
            created, updated, skipped, errors, feedbacks = MasterBulkUploadService.sync_from_path(db)
            mode = "sync"
        else:
            file_bytes = await file.read()
            created, updated, skipped, errors, feedbacks = MasterBulkUploadService.from_bytes(db, file_bytes)
            mode = "upload"
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("master_results_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    log.info("master_results", mode=mode, created=created, updated=updated, user=current_user.email)
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        feedbacks=feedbacks,
        message=f"Master Results ({mode}): {created} created, {updated} updated, {skipped} skipped",
    )


@router.post("/icp-oes", response_model=UploadResponse)
async def upload_icp_oes(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload an ICP-OES CSV file and ingest elemental data."""
    import sys  # noqa: PLC0415
    from types import ModuleType  # noqa: PLC0415
    if "frontend.config.variable_config" not in sys.modules:
        _stub = ModuleType("frontend.config.variable_config")
        sys.modules["frontend"] = sys.modules.get("frontend", ModuleType("frontend"))
        sys.modules["frontend.config"] = sys.modules.get("frontend.config", ModuleType("frontend.config"))
        sys.modules["frontend.config.variable_config"] = _stub
    _vc = sys.modules["frontend.config.variable_config"]
    if not hasattr(_vc, "ICP_FIXED_ELEMENT_FIELDS"):
        _vc.ICP_FIXED_ELEMENT_FIELDS = [
            "fe", "si", "mg", "ca", "ni", "cu", "mo", "zn", "mn", "cr",
            "co", "al", "sr", "y", "nb", "sb", "cs", "ba", "nd", "gd",
            "pt", "rh", "ir", "pd", "ru", "os", "tl",
        ]
    from backend.services.icp_service import ICPService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        processed_data, parse_errors = ICPService.parse_and_process_icp_file(file_bytes)
        if parse_errors and not processed_data:
            return UploadResponse(created=0, updated=0, skipped=0, errors=parse_errors,
                                  message="ICP parse failed")
        created_rows, updated_count, ingest_errors = ICPService.bulk_create_icp_results(db, processed_data)
        all_errors = parse_errors + ingest_errors
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("icp_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    new_count = len(created_rows) - updated_count
    log.info("icp_upload", created=new_count, updated=updated_count, user=current_user.email)
    return UploadResponse(
        created=new_count, updated=updated_count, skipped=0, errors=all_errors,
        message=f"ICP-OES: {new_count} created, {updated_count} updated",
    )


@router.post("/xrd-mineralogy", response_model=UploadResponse)
async def upload_xrd_mineralogy(
    file: UploadFile = File(...),
    overwrite: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload an XRD file — auto-detects Aeris, ActLabs, or Experiment+Timepoint format.

    When overwrite=True, all existing XRDPhase rows for each matching key
    (experiment+timepoint or sample) are deleted before the new phases are inserted.
    """
    from backend.services.bulk_uploads.xrd_upload import XRDAutoDetectService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors = XRDAutoDetectService.upload(
            db, file_bytes, overwrite=overwrite
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("xrd_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        message=f"XRD: {created} created, {updated} updated",
    )


@router.post("/timepoint-modifications", response_model=UploadResponse)
async def upload_timepoint_modifications(
    file: UploadFile = File(...),
    overwrite_existing: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Bulk-set brine modification descriptions on existing result rows."""
    from backend.services.bulk_uploads.timepoint_modifications import TimepointModificationsService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        updated, skipped, errors, feedbacks = TimepointModificationsService.bulk_set_from_bytes(
            db, file_bytes,
            overwrite_existing=overwrite_existing,
            modified_by=current_user.email,
        )
        if not errors:
            db.commit()
        else:
            db.rollback()
    except Exception as exc:
        db.rollback()
        log.error("timepoint_mod_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(
        created=0, updated=updated, skipped=skipped, errors=errors,
        feedbacks=feedbacks,
        message=f"Timepoint Modifications: {updated} updated, {skipped} skipped",
    )


@router.post("/rock-inventory", response_model=UploadResponse)
async def upload_rock_inventory(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload a Rock Inventory Excel file."""
    from backend.services.bulk_uploads.rock_inventory import RockInventoryService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, _images, skipped, errors, warnings = (
            RockInventoryService.bulk_upsert_samples(db, file_bytes, [])
        )
        if not errors:
            db.commit()
        else:
            db.rollback()
    except Exception as exc:
        db.rollback()
        log.error("rock_inventory_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        warnings=warnings,
        message=f"Rock Inventory: {created} created, {updated} updated",
    )


@router.post("/chemical-inventory", response_model=UploadResponse)
async def upload_chemical_inventory(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload a Chemical Inventory Excel file."""
    from backend.services.bulk_uploads.chemical_inventory import ChemicalInventoryService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors = ChemicalInventoryService.bulk_upsert_from_excel(
            db, file_bytes
        )
        if not errors:
            db.commit()
        else:
            db.rollback()
    except Exception as exc:
        db.rollback()
        log.error("chemical_inventory_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        message=f"Chemical Inventory: {created} created, {updated} updated",
    )


@router.post("/elemental-composition", response_model=UploadResponse)
async def upload_elemental_composition(
    file: UploadFile = File(...),
    default_unit: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """
    Upload Sample Chemical Composition (wide-format Excel).

    ``default_unit`` — when provided, any analyte column not already in the
    Analyte table is auto-created with this unit (e.g. "ppm", "%", "wt%").
    """
    from backend.services.bulk_uploads.actlabs_titration_data import ElementalCompositionService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
            db, file_bytes, default_unit=default_unit
        )
        if not errors:
            db.commit()
        else:
            db.rollback()
    except Exception as exc:
        db.rollback()
        log.error("elemental_composition_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        message=f"Elemental Composition: {created} created, {updated} updated",
    )


@router.post("/actlabs-rock", response_model=UploadResponse)
async def upload_actlabs_rock(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload an ActLabs Rock Analysis file (titration report)."""
    from backend.services.bulk_uploads.actlabs_titration_data import ActlabsRockTitrationService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        created, updated, skipped, errors = ActlabsRockTitrationService.import_excel(db, file_bytes)
        db.commit()
    except Exception as exc:
        db.rollback()
        log.error("actlabs_rock_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    return UploadResponse(
        created=created, updated=updated, skipped=skipped, errors=errors,
        message=f"ActLabs Rock: {created} created, {updated} updated",
    )


@router.post("/experiment-status", response_model=UploadResponse)
async def upload_experiment_status(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> UploadResponse:
    """Upload an Experiment Status Excel file (bulk ONGOING / COMPLETED changes)."""
    from backend.services.bulk_uploads.experiment_status import ExperimentStatusService  # noqa: PLC0415
    file_bytes = await file.read()
    try:
        preview = ExperimentStatusService.preview_status_changes_from_excel(db, file_bytes)
        if preview.errors:
            return UploadResponse(
                created=0, updated=0, skipped=len(preview.missing_ids),
                errors=preview.errors,
                message="Validation failed — no changes applied",
            )
        to_ongoing_ids = [item["experiment_id"] for item in preview.to_ongoing]
        reactor_map = {
            item["experiment_id"]: item["new_reactor_number"]
            for item in preview.to_ongoing
            if item.get("new_reactor_number") is not None
        }
        marked_ongoing, marked_completed, _reactor_updates, errors = (
            ExperimentStatusService.apply_status_changes(db, to_ongoing_ids, reactor_map)
        )
        if not errors:
            db.commit()
        else:
            db.rollback()
    except Exception as exc:
        db.rollback()
        log.error("experiment_status_upload_failed", error=str(exc))
        return UploadResponse(created=0, updated=0, skipped=0, errors=[str(exc)],
                              message="Upload failed")
    total = marked_ongoing + marked_completed
    return UploadResponse(
        created=0, updated=total, skipped=len(preview.missing_ids),
        errors=errors,
        feedbacks=[
            {"marked_ongoing": marked_ongoing, "marked_completed": marked_completed}
        ],
        message=(
            f"Status update: {marked_ongoing} → ONGOING, "
            f"{marked_completed} → COMPLETED, {len(preview.missing_ids)} not found"
        ),
    )


# ---------------------------------------------------------------------------
# Master Results config (Chunk D)
# ---------------------------------------------------------------------------

from pydantic import BaseModel as _PydanticBase  # noqa: E402


class MasterResultsConfigResponse(_PydanticBase):
    path: str | None


class MasterResultsConfigUpdate(_PydanticBase):
    path: str


_MASTER_RESULTS_CONFIG_KEY = "master_results_path"


@router.get("/master-results/config", response_model=MasterResultsConfigResponse)
def get_master_results_config(
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> MasterResultsConfigResponse:
    """Return the currently configured Master Results file path."""
    from database.models.app_config import AppConfig  # noqa: PLC0415
    cfg = db.query(AppConfig).filter_by(key=_MASTER_RESULTS_CONFIG_KEY).first()
    if cfg:
        return MasterResultsConfigResponse(path=cfg.value)
    from backend.config.settings import get_settings  # noqa: PLC0415
    return MasterResultsConfigResponse(path=get_settings().master_results_path)


@router.patch("/master-results/config", response_model=MasterResultsConfigResponse)
def update_master_results_config(
    body: MasterResultsConfigUpdate,
    db: Session = Depends(get_db),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> MasterResultsConfigResponse:
    """Set the Master Results file path after validating it resolves to a readable .xlsx file."""
    import os  # noqa: PLC0415
    from database.models.app_config import AppConfig  # noqa: PLC0415

    path = body.path
    if not os.path.isfile(path):
        raise HTTPException(status_code=422, detail=f"File not found: {path}")
    if not path.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=422, detail="Path must point to an .xlsx or .xls file")

    cfg = db.query(AppConfig).filter_by(key=_MASTER_RESULTS_CONFIG_KEY).first()
    if cfg:
        cfg.value = path
    else:
        db.add(AppConfig(key=_MASTER_RESULTS_CONFIG_KEY, value=path))
    db.commit()
    return MasterResultsConfigResponse(path=path)


# ---------------------------------------------------------------------------
# Template downloads
# ---------------------------------------------------------------------------

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_NO_TEMPLATE = {
    "master-results", "icp-oes", "aeris-xrd", "actlabs-rock", "pxrf",
}


def _simple_template(
    headers: list[str],
    required: set[str],
    sheet_title: str = "Template",
    example_row: list | None = None,
) -> bytes:
    """Generate a minimal Excel template with highlighted required columns."""
    import openpyxl  # noqa: PLC0415
    from openpyxl.styles import PatternFill, Font, Alignment  # noqa: PLC0415

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    req_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    opt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = req_fill if h in required else opt_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    if example_row:
        for col, val in enumerate(example_row, start=1):
            ws.cell(row=2, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_template_bytes(upload_type: str, mode: Optional[str] = None) -> bytes:
    if upload_type == "xrd-mineralogy":
        from backend.services.bulk_uploads.xrd_upload import XRDAutoDetectService  # noqa: PLC0415
        return XRDAutoDetectService.generate_template_bytes(mode=mode or "sample")

    if upload_type == "timepoint-modifications":
        from backend.services.bulk_uploads.timepoint_modifications import TimepointModificationsService  # noqa: PLC0415
        return TimepointModificationsService.generate_template_bytes()

    if upload_type == "scalar-results":
        return _simple_template(
            headers=[
                "Experiment ID", "Time (days)", "Description", "Date",
                "Gross Ammonium (mM)", "Sampling Vol (mL)", "Bkg Ammonium (mM)", "Bkg Exp ID",
                "H2 Conc (ppm)", "Gas Sample Vol (mL)", "Gas Pressure (MPa)",
                "Final pH", "Fe2+ Yield (%)", "Final DO (mg/L)",
                "Conductivity (mS/cm)", "Overwrite",
            ],
            required={"Experiment ID", "Time (days)"},
            example_row=["HPHT_001", 7.0, "Day 7 sample", None, 5.2, 2.0, 0.3, None,
                         120.0, 5.0, 0.5, 7.2, None, None, 12.5, "FALSE"],
        )

    if upload_type == "new-experiments":
        return _simple_template(
            headers=["experiment_id", "experiment_type", "date", "researcher",
                     "sample_id", "temperature_c", "description"],
            required={"experiment_id", "experiment_type"},
            example_row=["HPHT_072", "HPHT", None, "MH", "S001", 200.0, ""],
        )

    if upload_type == "rock-inventory":
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import PatternFill, Font, Alignment  # noqa: PLC0415

        headers = [
            "sample_id", "rock_classification", "state", "country",
            "locality", "latitude", "longitude", "description",
            "characterized", "pxrf_reading_no", "magnetic_susceptibility",
            "well_name", "core_lender", "core_interval_ft", "on_loan_return_date",
            "overwrite",
        ]
        required = {"sample_id"}
        example_row = [
            "S001", "Basalt", "BC", "Canada", "Vancouver Island",
            49.5, -125.0, "Fresh olivine basalt", "FALSE", "", "",
            "", "", "", "",
            "FALSE",
        ]
        req_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        opt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = req_fill if h in required else opt_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)
        for col, val in enumerate(example_row, start=1):
            ws.cell(row=2, column=col, value=val)

        ws_inst = wb.create_sheet("INSTRUCTIONS")
        ws_inst.column_dimensions["A"].width = 30
        ws_inst.column_dimensions["B"].width = 70
        instructions = [
            ("Column", "Notes"),
            ("sample_id", "REQUIRED. Unique sample identifier (e.g. S001, SROCK-042)."),
            ("rock_classification", "Rock type (e.g. Basalt, Dunite, Serpentinite)."),
            ("state", "Province or state."),
            ("country", "Country of origin."),
            ("locality", "Locality or formation name."),
            ("latitude", "Decimal degrees (e.g. 49.5)."),
            ("longitude", "Decimal degrees (e.g. -125.0)."),
            ("description", "Free-text sample description."),
            ("characterized", "TRUE or FALSE (default FALSE)."),
            (
                "pxrf_reading_no",
                "Comma-separated pXRF reading numbers. Creates ExternalAnalysis type 'pXRF' per reading.",
            ),
            (
                "magnetic_susceptibility",
                "Magnetic susceptibility value (units: 1x10\u207b\u00b3 SI). Leave blank if not measured.",
            ),
            (
                "well_name",
                "Well or borehole name (e.g. Tuscarora Project CT-3). Applies to core samples only.",
            ),
            (
                "core_lender",
                "Organization lending the core (e.g. Geologica). Applies to core samples only.",
            ),
            (
                "core_interval_ft",
                "Depth interval in feet as a string (e.g. 895'). Applies to core samples only.",
            ),
            (
                "on_loan_return_date",
                "Date core must be returned to lender (YYYY-MM-DD). Applies to core samples only.",
            ),
            ("overwrite", "TRUE clears and rewrites all optional fields for existing samples (default FALSE)."),
        ]
        for r_idx, (col_name, note) in enumerate(instructions, start=1):
            name_cell = ws_inst.cell(row=r_idx, column=1, value=col_name)
            note_cell = ws_inst.cell(row=r_idx, column=2, value=note)
            if r_idx == 1:
                name_cell.font = Font(bold=True)
                note_cell.font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if upload_type == "chemical-inventory":
        return _simple_template(
            headers=["name", "formula", "cas_number", "molecular_weight",
                     "density", "hazard_class", "supplier", "catalog_number", "notes"],
            required={"name"},
            example_row=["Magnesium hydroxide", "Mg(OH)2", "1309-42-8",
                         58.32, 2.34, None, "Sigma-Aldrich", "309907", None],
        )

    if upload_type == "elemental-composition":
        return _simple_template(
            headers=["sample_id", "SiO2", "Al2O3", "Fe2O3", "MgO", "CaO", "Na2O"],
            required={"sample_id"},
            example_row=["S001", 47.2, 13.5, 11.0, 8.4, 10.2, 2.6],
        )

    if upload_type == "experiment-status":
        return _simple_template(
            headers=["experiment_id", "reactor_number"],
            required={"experiment_id"},
            example_row=["HPHT_072", 3],
        )

    raise ValueError(f"Unknown template type: {upload_type}")


@router.get("/templates/{upload_type}")
async def download_template(
    upload_type: str,
    mode: Optional[str] = Query(None, description="Template mode (e.g. 'experiment' for XRD experiment+timepoint format)"),
    current_user: FirebaseUser = Depends(verify_firebase_token),
) -> StreamingResponse:
    """Download an Excel upload template for the specified upload type."""
    if upload_type in _NO_TEMPLATE:
        raise HTTPException(
            status_code=404,
            detail=f"No template available for '{upload_type}'. "
                   "This upload type uses instrument exports or fixed-format files.",
        )
    try:
        template_bytes = _get_template_bytes(upload_type, mode=mode)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        log.error("template_generation_failed", upload_type=upload_type, error=str(exc))
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {exc}")

    filename = f"{upload_type}-template.xlsx"
    return StreamingResponse(
        io.BytesIO(template_bytes),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
