"""Unified XRD upload: auto-detects Aeris, ActLabs, or Experiment+Timepoint format."""
from __future__ import annotations

import io
import re
from typing import List, Optional, Tuple

import pandas as pd
from sqlalchemy.orm import Session

from database.models import XRDPhase

# Aeris Sample ID pattern: DATE_ExperimentID-dDAYS_SCAN
# e.g. 20260218_HPHT070-d19_02
_AERIS_SAMPLE_RE = re.compile(r"^\d{8}_.+?-d\d+_\d+$")

_EXP_COL_VARIANTS = {"experiment_id", "experiment id"}
_TIME_COL_VARIANTS = {"time (days)", "time_days", "duration (days)", "time(days)"}
_DATE_COL_VARIANTS = {"date", "measurement_date", "measurement date"}
_SAMPLE_COL_VARIANTS = {"sample_id", "sample id"}


def _detect_format(file_bytes: bytes) -> str | None:
    """
    Inspect the first sheet to detect XRD format.

    Detection order:
    1. 'experiment-timepoint' — file has an Experiment ID column AND a Time (days) column
    2. 'aeris'               — sample_id column with Aeris-regex values
    3. 'actlabs'             — sample_id column with plain sample identifiers
    Returns None if no format can be identified.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), nrows=6)
    except Exception:
        return None

    cols_lower = [str(c).strip().lower() for c in df.columns]

    # Priority 1: explicit Experiment ID + time columns
    has_exp = any(c in _EXP_COL_VARIANTS for c in cols_lower)
    has_time = any(c in _TIME_COL_VARIANTS for c in cols_lower)
    if has_exp and has_time:
        return "experiment-timepoint"

    # Priority 2/3: sample_id column present
    sample_col_idx = next(
        (i for i, c in enumerate(cols_lower) if c in _SAMPLE_COL_VARIANTS), None
    )
    if sample_col_idx is None:
        return None

    orig_cols = [str(c).strip() for c in df.columns]
    sample_col = orig_cols[sample_col_idx]

    for val in df[sample_col].dropna().head(5):
        if _AERIS_SAMPLE_RE.match(str(val).strip()):
            return "aeris"

    return "actlabs"


def _find_experiment(db: Session, exp_id_raw: str):
    """Delimiter-insensitive experiment lookup (reused from aeris_xrd logic)."""
    from sqlalchemy import func  # noqa: PLC0415
    from database import Experiment  # noqa: PLC0415

    norm = "".join(ch for ch in exp_id_raw.lower() if ch not in ("-", "_", " "))
    return (
        db.query(Experiment)
        .filter(
            func.lower(
                func.replace(
                    func.replace(
                        func.replace(Experiment.experiment_id, "-", ""),
                        "_", "",
                    ),
                    " ", "",
                )
            )
            == norm
        )
        .first()
    )


def _clean_mineral_name(col: str) -> str:
    """Strip trailing [%] or (%) and whitespace from a column header."""
    col = re.sub(r"\s*[\[\(]%[\]\)]\s*$", "", col)
    return col.strip()


def _parse_experiment_timepoint(
    db: Session, file_bytes: bytes, overwrite: bool = False
) -> Tuple[int, int, int, List[str]]:
    """
    Parse a wide-format XRD file with explicit Experiment ID + Time (days) columns.

    Each row is one experiment at one timepoint; all remaining columns are mineral phases.
    When overwrite=False: upserts XRDPhase rows keyed on (experiment_id, time_post_reaction_days, mineral_name).
    When overwrite=True: deletes ALL existing XRDPhase rows for each (experiment_id, time_post_reaction_days)
    key before inserting the new set, so only the uploaded phases survive.

    Returns (created, updated, skipped, errors).
    """
    created = updated = skipped = 0
    errors: List[str] = []

    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception:
        try:
            df = pd.read_csv(io.BytesIO(file_bytes))
        except Exception as e:
            return 0, 0, 0, [f"Failed to read file: {e}"]

    cols_normalized = [str(c).strip() for c in df.columns]
    df.columns = cols_normalized
    cols_lower = [c.lower() for c in cols_normalized]

    exp_col = next(
        (cols_normalized[i] for i, c in enumerate(cols_lower) if c in _EXP_COL_VARIANTS),
        None,
    )
    if not exp_col:
        return 0, 0, 0, ["No 'Experiment ID' column found."]

    time_col = next(
        (cols_normalized[i] for i, c in enumerate(cols_lower) if c in _TIME_COL_VARIANTS),
        None,
    )
    if not time_col:
        return 0, 0, 0, ["No 'Time (days)' column found."]

    date_col = next(
        (cols_normalized[i] for i, c in enumerate(cols_lower) if c in _DATE_COL_VARIANTS),
        None,
    )

    identity_cols = {exp_col.lower(), time_col.lower()}
    if date_col:
        identity_cols.add(date_col.lower())
    mineral_cols = [c for c in cols_normalized if c.lower() not in identity_cols]
    if not mineral_cols:
        return 0, 0, 0, ["No mineral phase columns detected."]

    exp_cache: dict[str, Optional[object]] = {}
    cleared_keys: set[tuple[str, float]] = set()

    for idx, row in df.iterrows():
        row_num = idx + 2
        exp_id_raw = str(row.get(exp_col) or "").strip()
        if not exp_id_raw:
            skipped += 1
            continue

        time_raw = row.get(time_col)
        if time_raw is None or (isinstance(time_raw, float) and pd.isna(time_raw)):
            skipped += 1
            continue

        try:
            time_days = float(time_raw)
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: invalid Time (days) '{time_raw}'")
            continue

        if exp_id_raw not in exp_cache:
            exp_cache[exp_id_raw] = _find_experiment(db, exp_id_raw)
        experiment = exp_cache[exp_id_raw]

        if experiment is None:
            errors.append(f"Row {row_num}: experiment '{exp_id_raw}' not found.")
            continue

        measurement_date = None
        if date_col:
            date_raw = row.get(date_col)
            if date_raw is not None and not (isinstance(date_raw, float) and pd.isna(date_raw)):
                import datetime  # noqa: PLC0415
                if isinstance(date_raw, (datetime.date, datetime.datetime)):
                    measurement_date = date_raw
                else:
                    try:
                        measurement_date = datetime.datetime.strptime(str(date_raw).strip(), "%Y-%m-%d")
                    except ValueError:
                        pass  # Unrecognised format — leave as None

        # When overwrite=True, delete all existing phases for this (experiment, timepoint)
        # once per unique key within this file before inserting the new set.
        if overwrite:
            clear_key = (experiment.experiment_id, time_days)
            if clear_key not in cleared_keys:
                (
                    db.query(XRDPhase)
                    .filter(
                        XRDPhase.experiment_id == experiment.experiment_id,
                        XRDPhase.time_post_reaction_days == time_days,
                    )
                    .delete(synchronize_session=False)
                )
                db.flush()
                cleared_keys.add(clear_key)

        for mcol in mineral_cols:
            raw_val = row.get(mcol)
            if raw_val is None or (isinstance(raw_val, float) and pd.isna(raw_val)):
                continue
            try:
                amount_val = float(raw_val)
            except (ValueError, TypeError):
                continue

            mineral_name = _clean_mineral_name(mcol)

            if overwrite:
                # After clearing, always insert fresh
                db.add(XRDPhase(
                    experiment_fk=experiment.id,
                    experiment_id=experiment.experiment_id,
                    time_post_reaction_days=time_days,
                    mineral_name=mineral_name,
                    amount=amount_val,
                    measurement_date=measurement_date,
                ))
                created += 1
            else:
                phase = (
                    db.query(XRDPhase)
                    .filter(
                        XRDPhase.experiment_id == experiment.experiment_id,
                        XRDPhase.time_post_reaction_days == time_days,
                        XRDPhase.mineral_name == mineral_name,
                    )
                    .first()
                )

                if phase:
                    phase.amount = amount_val
                    phase.experiment_fk = experiment.id
                    if measurement_date is not None:
                        phase.measurement_date = measurement_date
                    updated += 1
                else:
                    db.add(XRDPhase(
                        experiment_fk=experiment.id,
                        experiment_id=experiment.experiment_id,
                        time_post_reaction_days=time_days,
                        mineral_name=mineral_name,
                        amount=amount_val,
                        measurement_date=measurement_date,
                    ))
                    created += 1

    return created, updated, skipped, errors


class XRDAutoDetectService:
    @staticmethod
    def upload(db: Session, file_bytes: bytes, overwrite: bool = False) -> Tuple[int, int, int, List[str]]:
        """
        Auto-detect XRD file format and delegate to the appropriate parser.

        Formats supported:
        - experiment-timepoint: 'Experiment ID' + 'Time (days)' columns (user-created)
        - aeris:   Aeris instrument export (Sample ID values like 20260218_HPHT070-d19_02)
        - actlabs: ActLabs report (plain sample_id column)

        The overwrite parameter is currently only applied to the experiment-timepoint format.
        For experiment-timepoint: when overwrite=True, all existing XRDPhase rows for each
        (experiment_id, time_post_reaction_days) key are deleted before the new set is inserted.

        Returns (created, updated, skipped, errors).
        """
        fmt = _detect_format(file_bytes)

        if fmt == "experiment-timepoint":
            return _parse_experiment_timepoint(db, file_bytes, overwrite=overwrite)

        if fmt == "aeris":
            from backend.services.bulk_uploads.aeris_xrd import AerisXRDUploadService  # noqa: PLC0415
            return AerisXRDUploadService.bulk_upsert_from_excel(db, file_bytes)

        if fmt == "actlabs":
            from backend.services.bulk_uploads.actlabs_xrd_report import XRDUploadService  # noqa: PLC0415
            (
                created_ext, updated_ext,
                created_json, updated_json,
                created_phase, updated_phase,
                skipped, errors,
            ) = XRDUploadService.bulk_upsert_from_excel(db, file_bytes)
            created = created_phase + created_ext
            updated = updated_phase + updated_ext
            return created, updated, skipped, errors

        return 0, 0, 0, [
            "Unable to detect XRD file format. Expected one of:\n"
            "  (1) Experiment+Timepoint: columns 'Experiment ID' and 'Time (days)'\n"
            "  (2) Aeris instrument export: 'Sample ID' values like '20260218_HPHT070-d19_02'\n"
            "  (3) ActLabs format: 'sample_id' column with plain sample identifiers"
        ]

    @staticmethod
    def generate_template_bytes(mode: str = "sample") -> bytes:
        """
        Return a downloadable XRD template as Excel bytes.

        mode='sample'      — ActLabs-style: sample_id + mineral columns
        mode='experiment'  — Experiment+Timepoint: Experiment ID + Time (days) + mineral columns
        """
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import PatternFill, Font, Alignment  # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active

        req_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        opt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        if mode == "experiment":
            ws.title = "XRD Experiment-Timepoint"
            headers = ["Experiment ID", "Time (days)", "Date", "Quartz", "Calcite", "Dolomite", "Olivine", "Serpentine"]
            required = {"Experiment ID", "Time (days)"}
            example = ["HPHT_001", 7.0, "2026-03-19", 45.2, 20.1, 15.0, 10.5, 9.2]
        else:
            ws.title = "XRD Sample-Based"
            headers = ["sample_id", "Quartz", "Calcite", "Dolomite", "Feldspar", "Pyrite", "Other"]
            required = {"sample_id"}
            example = ["S001", 45.2, 20.1, 15.0, 10.5, 5.2, 4.0]

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = req_fill if h in required else opt_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 18)

        for col, val in enumerate(example, start=1):
            ws.cell(row=2, column=col, value=val)

        instr = wb.create_sheet("INSTRUCTIONS")
        if mode == "experiment":
            rows = [
                ["XRD Mineralogy — Experiment + Timepoint Format"],
                [""],
                ["Use this format when uploading post-reaction XRD data for specific experiments."],
                [""],
                ["Column", "Description"],
                ["Experiment ID (required)", "Must match an existing experiment (delimiter-insensitive, e.g. HPHT_001 or HPHT001)."],
                ["Time (days) (required)", "Days post-reaction as a number (e.g. 7, 14.5). Use 0 for pre-reaction baseline."],
                ["Date (optional)", "Measurement date in YYYY-MM-DD format. Stored as measurement_date on the XRDPhase record."],
                ["[mineral name]", "Mineral weight percent (0–100). Add/remove columns as needed. Blank cells are skipped."],
                [""],
                ["Notes:"],
                ["- Column headers become mineral phase names in the database."],
                ["- Uploading again with the same Experiment ID + Time (days) + mineral name updates the existing value."],
                ["- Column headers may include trailing % indicators like 'Quartz (%)' — these are stripped automatically."],
            ]
        else:
            rows = [
                ["XRD Mineralogy — Sample-Based Format"],
                [""],
                ["Use this format for sample characterization data (not tied to a specific timepoint)."],
                [""],
                ["Column", "Description"],
                ["sample_id (required)", "Must match an existing sample in the database."],
                ["[mineral name]", "Mineral weight percent (0–100). Add/remove columns as needed. Blank cells are skipped."],
                [""],
                ["Format auto-detection:"],
                ["  Experiment+Timepoint — columns named 'Experiment ID' and 'Time (days)'"],
                ["  Aeris instrument export — 'Sample ID' values like '20260218_HPHT070-d19_02'"],
                ["  Sample-based (this template) — 'sample_id' column with plain identifiers"],
                [""],
                ["Notes:"],
                ["- Column headers become mineral phase names in the database."],
                ["- Column headers may include trailing % indicators — these are stripped automatically."],
            ]

        for r in rows:
            instr.append(r)
        instr.column_dimensions["A"].width = 35
        instr.column_dimensions["B"].width = 75

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
