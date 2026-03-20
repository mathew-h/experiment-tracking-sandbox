"""Bulk-set brine_modification_description on ExperimentalResults rows."""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from sqlalchemy.orm import Session

from database import Experiment, ExperimentalResults, ModificationsLog
from backend.services.result_merge_utils import find_timepoint_candidates

# Column aliases accepted in the upload file
_EXPERIMENT_ID_ALIASES = {"experiment_id", "experiment id", "exp_id", "exp id"}
_TIME_POINT_ALIASES = {"time_point", "time (days)", "time(days)", "duration (days)", "days"}
_MODIFICATION_ALIASES = {
    "modification_description", "experiment_modification", "modification", "description",
    "brine_modification_description",
}
_OVERWRITE_ALIASES = {"overwrite_existing", "overwrite"}


def _resolve_col(columns: List[str], aliases: set[str]) -> Optional[str]:
    for c in columns:
        if c.strip().lower() in aliases:
            return c
    return None


def _parse_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in {"true", "yes", "1", "y"}
    return False


class TimepointModificationsService:
    @staticmethod
    def bulk_set_from_bytes(
        db: Session,
        file_bytes: bytes,
        overwrite_existing: bool = False,
        modified_by: str = "bulk_upload",
    ) -> Tuple[int, int, List[str], List[Dict[str, Any]]]:
        """
        Read an Excel/CSV file and set brine_modification_description on
        matching ExperimentalResults rows.

        The model @validates('brine_modification_description') automatically
        syncs has_brine_modification.

        Returns (updated, skipped, errors, feedbacks).
        """
        updated = skipped = 0
        errors: List[str] = []
        feedbacks: List[Dict[str, Any]] = []

        # --- Load file -------------------------------------------------------
        try:
            if file_bytes[:2] in (b"PK", b"\xd0\xcf"):  # xlsx / xls
                df = pd.read_excel(io.BytesIO(file_bytes))
            else:
                df = pd.read_csv(io.BytesIO(file_bytes))
        except Exception as exc:
            try:
                df = pd.read_excel(io.BytesIO(file_bytes))
            except Exception:
                return 0, 0, [f"Failed to read file: {exc}"], []

        df.columns = [str(c).strip() for c in df.columns]
        cols = list(df.columns)

        exp_col = _resolve_col(cols, _EXPERIMENT_ID_ALIASES)
        tp_col = _resolve_col(cols, _TIME_POINT_ALIASES)
        mod_col = _resolve_col(cols, _MODIFICATION_ALIASES)
        ow_col = _resolve_col(cols, _OVERWRITE_ALIASES)

        missing = []
        if not exp_col:
            missing.append("experiment_id")
        if not tp_col:
            missing.append("time_point")
        if not mod_col:
            missing.append("modification_description")
        if missing:
            return 0, 0, [f"Missing required columns: {', '.join(missing)}"], []

        # --- Pre-scan for duplicates within the file -------------------------
        seen: set[tuple[str, float]] = set()
        duplicates: List[str] = []
        for idx, row in df.iterrows():
            exp_id = str(row.get(exp_col) or "").strip()
            tp_raw = row.get(tp_col)
            if not exp_id or tp_raw is None:
                continue
            try:
                tp = float(tp_raw)
            except Exception:
                continue
            key = (exp_id, tp)
            if key in seen:
                duplicates.append(f"Row {idx + 2}: duplicate ({exp_id}, {tp})")
            seen.add(key)

        if duplicates:
            return 0, 0, [
                "File contains duplicate (experiment_id, time_point) pairs. "
                "Resolve duplicates and re-upload.",
                *duplicates,
            ], []

        # --- Process rows ----------------------------------------------------
        for idx, row in df.iterrows():
            row_num = idx + 2  # 1-based, header on row 1
            exp_id = str(row.get(exp_col) or "").strip()
            tp_raw = row.get(tp_col)
            modification = str(row.get(mod_col) or "").strip()

            if not exp_id:
                skipped += 1
                continue
            if tp_raw is None or (isinstance(tp_raw, float) and pd.isna(tp_raw)):
                skipped += 1
                continue
            try:
                tp = float(tp_raw)
            except Exception:
                errors.append(f"Row {row_num}: invalid time_point '{tp_raw}'")
                continue

            # Per-row overwrite flag (file column) takes precedence over global param
            row_overwrite = overwrite_existing
            if ow_col:
                ow_val = row.get(ow_col)
                if ow_val is not None and not (isinstance(ow_val, float) and pd.isna(ow_val)):
                    row_overwrite = _parse_bool(ow_val)

            # --- Resolve experiment ---
            experiment = (
                db.query(Experiment)
                .filter(Experiment.experiment_id == exp_id)
                .first()
            )
            if not experiment:
                errors.append(f"Row {row_num}: experiment '{exp_id}' not found")
                continue

            # --- Find matching timepoint row ---
            candidates = find_timepoint_candidates(db, experiment.id, tp)
            if not candidates:
                errors.append(
                    f"Row {row_num}: no result row found for experiment='{exp_id}' "
                    f"time_point={tp}"
                )
                continue

            target: ExperimentalResults = candidates[0]

            # --- Overwrite guard ---
            if target.brine_modification_description and not row_overwrite:
                skipped += 1
                feedbacks.append({
                    "row": row_num,
                    "experiment_id": exp_id,
                    "time_point": tp,
                    "status": "skipped",
                    "reason": "modification already set; pass overwrite_existing=true to replace",
                })
                continue

            # --- Apply ---
            old_val = target.brine_modification_description
            target.brine_modification_description = modification or None

            # Audit log
            db.add(ModificationsLog(
                experiment_id=exp_id,
                experiment_fk=experiment.id,
                modified_by=modified_by,
                modification_type="update",
                modified_table="experimental_results",
                old_values={"brine_modification_description": old_val},
                new_values={"brine_modification_description": modification or None},
            ))
            updated += 1
            feedbacks.append({
                "row": row_num,
                "experiment_id": exp_id,
                "time_point": tp,
                "status": "updated",
                "result_id": target.id,
            })

        return updated, skipped, errors, feedbacks

    @staticmethod
    def generate_template_bytes() -> bytes:
        """Return a downloadable Excel template for timepoint modifications."""
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import PatternFill, Font, Alignment  # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"

        headers = [
            "experiment_id",
            "time_point",
            "modification_description",
            "overwrite_existing",
        ]
        required = {"experiment_id", "time_point", "modification_description"}
        req_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        opt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = req_fill if h in required else opt_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 28

        # Example row
        ws.cell(row=2, column=1, value="HPHT_001")
        ws.cell(row=2, column=2, value=7.0)
        ws.cell(row=2, column=3, value="Added 5 g Mg(OH)2 after sampling")
        ws.cell(row=2, column=4, value="FALSE")

        # Instructions sheet
        instr = wb.create_sheet("INSTRUCTIONS")
        rows = [
            ["Timepoint Modifications — Upload Instructions"],
            [""],
            ["Column", "Required", "Description"],
            ["experiment_id", "Yes",
             "Experiment identifier (e.g. HPHT_001). Must exist in the database."],
            ["time_point", "Yes",
             "Time post-reaction in days (float). Matched with ±0.0001 day tolerance."],
            ["modification_description", "Yes",
             "Free-text description of the modification at this timepoint."],
            ["overwrite_existing", "No",
             "Set TRUE to overwrite an existing modification. Default: FALSE."],
            [""],
            ["Notes:"],
            ["- Yellow headers are required. Green headers are optional."],
            ["- If a matching result row already has a modification description,"],
            ["  the row will be skipped unless overwrite_existing=TRUE."],
            ["- Duplicate (experiment_id, time_point) pairs in a single file are rejected."],
        ]
        for r in rows:
            instr.append(r)
        instr.column_dimensions["A"].width = 30
        instr.column_dimensions["B"].width = 12
        instr.column_dimensions["C"].width = 70

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
