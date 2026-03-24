"""Tests for ElementalCompositionService."""
from __future__ import annotations

import pytest
from sqlalchemy.orm import Session

from database import Analyte, ElementalAnalysis, SampleInfo
from database.models.analysis import ExternalAnalysis
from backend.services.bulk_uploads.actlabs_titration_data import ElementalCompositionService

from .excel_helpers import make_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _seed_sample(db: Session, sample_id: str) -> SampleInfo:
    sample = SampleInfo(sample_id=sample_id)
    db.add(sample)
    db.flush()
    return sample


def _seed_analyte(db: Session, symbol: str, unit: str = "wt%") -> Analyte:
    analyte = Analyte(analyte_symbol=symbol, unit=unit)
    db.add(analyte)
    db.flush()
    return analyte


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_creates_elemental_analysis_for_known_analyte(db_session: Session):
    """Valid file with a known analyte creates ElementalAnalysis rows."""
    _seed_sample(db_session, "S_ELEM001")
    _seed_analyte(db_session, "SiO2")

    xlsx = make_excel(
        ["sample_id", "SiO2"],
        [["S_ELEM001", 47.2]],
    )
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx
    )

    assert errors == [], f"Unexpected errors: {errors}"
    assert created == 1

    db_session.flush()
    analysis = (
        db_session.query(ElementalAnalysis)
        .join(Analyte, Analyte.id == ElementalAnalysis.analyte_id)
        .filter(Analyte.analyte_symbol == "SiO2")
        .first()
    )
    assert analysis is not None
    assert analysis.analyte_composition == pytest.approx(47.2)
    assert analysis.sample_id == "S_ELEM001"


def test_auto_creates_external_analysis_stub(db_session: Session):
    """Service auto-creates a 'Bulk Elemental Composition' ExternalAnalysis stub."""
    _seed_sample(db_session, "S_ELEM_STUB")
    _seed_analyte(db_session, "CaO")

    xlsx = make_excel(["sample_id", "CaO"], [["S_ELEM_STUB", 5.1]])
    ElementalCompositionService.bulk_upsert_wide_from_excel(db_session, xlsx)

    db_session.flush()
    stub = (
        db_session.query(ExternalAnalysis)
        .filter(
            ExternalAnalysis.sample_id == "S_ELEM_STUB",
            ExternalAnalysis.analysis_type == "Bulk Elemental Composition",
        )
        .first()
    )
    assert stub is not None


def test_unknown_analyte_without_default_unit_skipped(db_session: Session):
    """Unknown analyte header with no default_unit → silently skipped."""
    _seed_sample(db_session, "S_ELEM002")

    xlsx = make_excel(
        ["sample_id", "UnknownOxide"],
        [["S_ELEM002", 5.0]],
    )
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx, default_unit=None
    )

    assert errors == []
    assert created == 0
    count = (
        db_session.query(ElementalAnalysis)
        .join(Analyte, Analyte.id == ElementalAnalysis.analyte_id)
        .filter(Analyte.analyte_symbol == "UnknownOxide")
        .count()
    )
    assert count == 0


def test_unknown_analyte_with_default_unit_auto_created(db_session: Session):
    """Unknown analyte header with default_unit → auto-created Analyte and data stored."""
    _seed_sample(db_session, "S_ELEM003")

    xlsx = make_excel(
        ["sample_id", "NewElement"],
        [["S_ELEM003", 12.5]],
    )
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx, default_unit="ppm"
    )

    assert errors == [], f"Unexpected errors: {errors}"
    assert created == 1

    db_session.flush()
    analyte = (
        db_session.query(Analyte)
        .filter(Analyte.analyte_symbol == "NewElement")
        .first()
    )
    assert analyte is not None
    assert analyte.unit == "ppm"


def test_unknown_sample_id_recorded_as_error(db_session: Session):
    """Row with a sample_id not in the DB is recorded as an error."""
    _seed_analyte(db_session, "Al2O3")

    xlsx = make_excel(
        ["sample_id", "Al2O3"],
        [["S_NOTEXIST", 13.5]],
    )
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx
    )

    assert created == 0
    assert any("S_NOTEXIST" in e for e in errors)


def test_updates_existing_elemental_analysis(db_session: Session):
    """Re-uploading the same sample+analyte with overwrite=True updates, not creates."""
    _seed_sample(db_session, "S_ELEM004")
    _seed_analyte(db_session, "Fe2O3")

    xlsx1 = make_excel(["sample_id", "Fe2O3"], [["S_ELEM004", 10.0]])
    ElementalCompositionService.bulk_upsert_wide_from_excel(db_session, xlsx1)
    db_session.flush()

    xlsx2 = make_excel(["sample_id", "Fe2O3"], [["S_ELEM004", 15.0]])
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx2, overwrite=True
    )

    assert errors == [], f"Unexpected errors: {errors}"
    assert updated == 1
    assert created == 0


def test_no_sample_id_column_returns_error(db_session: Session):
    """File without a 'sample_id' column returns an error immediately."""
    xlsx = make_excel(
        ["SiO2", "Al2O3"],
        [[47.2, 13.5]],
    )
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx
    )

    assert created == 0
    assert any("sample_id" in e.lower() for e in errors)


def test_blank_sample_id_rows_skipped(db_session: Session):
    """Rows with blank sample_id are skipped."""
    _seed_analyte(db_session, "MgO")

    xlsx = make_excel(
        ["sample_id", "MgO"],
        [["   ", 8.5]],  # spaces-only → .strip() → "" → skipped
    )
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx
    )

    assert created == 0
    assert skipped == 1


def test_skip_existing_overwrite_false(db_session: Session):
    """overwrite=False (default): re-uploading an existing (sample_id, analyte) preserves original value."""
    _seed_sample(db_session, "S_SKIP01")
    _seed_analyte(db_session, "TiO2")

    xlsx1 = make_excel(["sample_id", "TiO2"], [["S_SKIP01", 3.5]])
    ElementalCompositionService.bulk_upsert_wide_from_excel(db_session, xlsx1)
    db_session.flush()

    xlsx2 = make_excel(["sample_id", "TiO2"], [["S_SKIP01", 9.9]])
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx2, overwrite=False
    )

    assert errors == []
    assert created == 0
    assert updated == 0

    row = (
        db_session.query(ElementalAnalysis)
        .join(Analyte, Analyte.id == ElementalAnalysis.analyte_id)
        .filter(Analyte.analyte_symbol == "TiO2", ElementalAnalysis.sample_id == "S_SKIP01")
        .first()
    )
    assert row is not None
    assert row.analyte_composition == pytest.approx(3.5), "Existing value must be preserved when overwrite=False"


def test_overwrite_existing_overwrite_true(db_session: Session):
    """overwrite=True: re-uploading an existing (sample_id, analyte) replaces the value."""
    _seed_sample(db_session, "S_OVR01")
    _seed_analyte(db_session, "K2O")

    xlsx1 = make_excel(["sample_id", "K2O"], [["S_OVR01", 0.5]])
    ElementalCompositionService.bulk_upsert_wide_from_excel(db_session, xlsx1)
    db_session.flush()

    xlsx2 = make_excel(["sample_id", "K2O"], [["S_OVR01", 1.8]])
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx2, overwrite=True
    )

    assert errors == []
    assert updated == 1
    assert created == 0

    row = (
        db_session.query(ElementalAnalysis)
        .join(Analyte, Analyte.id == ElementalAnalysis.analyte_id)
        .filter(Analyte.analyte_symbol == "K2O", ElementalAnalysis.sample_id == "S_OVR01")
        .first()
    )
    assert row.analyte_composition == pytest.approx(1.8)


def test_null_cell_does_not_clear_existing_value(db_session: Session):
    """Null/blank cell in the upload file never clears an existing value, even with overwrite=True."""
    _seed_sample(db_session, "S_NULL01")
    _seed_analyte(db_session, "P2O5")

    xlsx1 = make_excel(["sample_id", "P2O5"], [["S_NULL01", 0.08]])
    ElementalCompositionService.bulk_upsert_wide_from_excel(db_session, xlsx1)
    db_session.flush()

    xlsx2 = make_excel(["sample_id", "P2O5"], [["S_NULL01", None]])
    created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(
        db_session, xlsx2, overwrite=True
    )

    assert errors == []
    row = (
        db_session.query(ElementalAnalysis)
        .join(Analyte, Analyte.id == ElementalAnalysis.analyte_id)
        .filter(Analyte.analyte_symbol == "P2O5", ElementalAnalysis.sample_id == "S_NULL01")
        .first()
    )
    assert row is not None
    assert row.analyte_composition == pytest.approx(0.08), "Null cell must never clear existing value"


# ---------------------------------------------------------------------------
# ActlabsRockTitrationService overwrite tests
# ---------------------------------------------------------------------------

def _make_actlabs_xlsx(sample_id: str, values: dict) -> bytes:
    """Build a minimal ActLabs-format xlsx with one data row."""
    import io
    import openpyxl
    symbols = list(values.keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Report Number", "12345"])
    ws.append(["Report Date", "2026-01-01"])
    ws.append(["sample_id"] + symbols)
    ws.append([None] + ["ppm"] * len(symbols))
    ws.append([None] + ["0.01"] * len(symbols))
    ws.append(["Analysis Method"] + ["FUS-ICP"] * len(symbols))
    ws.append([sample_id] + [values[s] for s in symbols])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_actlabs_skip_existing_overwrite_false(db_session: Session):
    """ActlabsRockTitrationService: overwrite=False preserves existing values."""
    from backend.services.bulk_uploads.actlabs_titration_data import ActlabsRockTitrationService

    sample = SampleInfo(sample_id="S_ACT_SKIP")
    db_session.add(sample)
    db_session.flush()

    xlsx1 = _make_actlabs_xlsx("S_ACT_SKIP", {"Fe": 55000.0})
    ActlabsRockTitrationService.import_excel(db_session, xlsx1)
    db_session.flush()

    xlsx2 = _make_actlabs_xlsx("S_ACT_SKIP", {"Fe": 99999.0})
    created, updated, skipped, errors = ActlabsRockTitrationService.import_excel(
        db_session, xlsx2, overwrite=False
    )

    assert errors == []
    assert updated == 0

    analyte = db_session.query(Analyte).filter(Analyte.analyte_symbol.ilike("Fe")).first()
    row = db_session.query(ElementalAnalysis).filter(
        ElementalAnalysis.sample_id == "S_ACT_SKIP",
        ElementalAnalysis.analyte_id == analyte.id,
    ).first()
    assert row.analyte_composition == pytest.approx(55000.0), "Existing value must be preserved when overwrite=False"


def test_actlabs_overwrite_existing_overwrite_true(db_session: Session):
    """ActlabsRockTitrationService: overwrite=True replaces existing values."""
    from backend.services.bulk_uploads.actlabs_titration_data import ActlabsRockTitrationService

    sample = SampleInfo(sample_id="S_ACT_OVR")
    db_session.add(sample)
    db_session.flush()

    xlsx1 = _make_actlabs_xlsx("S_ACT_OVR", {"Mg": 12000.0})
    ActlabsRockTitrationService.import_excel(db_session, xlsx1)
    db_session.flush()

    xlsx2 = _make_actlabs_xlsx("S_ACT_OVR", {"Mg": 88888.0})
    created, updated, skipped, errors = ActlabsRockTitrationService.import_excel(
        db_session, xlsx2, overwrite=True
    )

    assert errors == []
    assert updated == 1

    analyte = db_session.query(Analyte).filter(Analyte.analyte_symbol.ilike("Mg")).first()
    row = db_session.query(ElementalAnalysis).filter(
        ElementalAnalysis.sample_id == "S_ACT_OVR",
        ElementalAnalysis.analyte_id == analyte.id,
    ).first()
    assert row.analyte_composition == pytest.approx(88888.0)


def test_actlabs_null_cell_does_not_clear_existing(db_session: Session):
    """ActlabsRockTitrationService: nd/blank cells never clear existing values, even with overwrite=True."""
    import io
    import openpyxl
    from backend.services.bulk_uploads.actlabs_titration_data import ActlabsRockTitrationService

    sample = SampleInfo(sample_id="S_ACT_NULL")
    db_session.add(sample)
    db_session.flush()

    xlsx1 = _make_actlabs_xlsx("S_ACT_NULL", {"Ni": 500.0})
    ActlabsRockTitrationService.import_excel(db_session, xlsx1)
    db_session.flush()

    # Second upload has "nd" (non-detect) for Ni — must not clear the existing value
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Report Number", "99999"])
    ws.append(["Report Date", "2026-01-01"])
    ws.append(["sample_id", "Ni"])
    ws.append([None, "ppm"])
    ws.append([None, "0.01"])
    ws.append(["Analysis Method", "FUS-ICP"])
    ws.append(["S_ACT_NULL", "nd"])
    buf = io.BytesIO()
    wb.save(buf)

    created, updated, skipped, errors = ActlabsRockTitrationService.import_excel(
        db_session, buf.getvalue(), overwrite=True
    )

    assert errors == []
    analyte = db_session.query(Analyte).filter(Analyte.analyte_symbol.ilike("Ni")).first()
    row = db_session.query(ElementalAnalysis).filter(
        ElementalAnalysis.sample_id == "S_ACT_NULL",
        ElementalAnalysis.analyte_id == analyte.id,
    ).first()
    assert row is not None
    assert row.analyte_composition == pytest.approx(500.0), "nd cell must never clear existing value"


def test_actlabs_import_sets_external_analysis_id(db_session: Session):
    """ActlabsRockTitrationService.import_excel must link every ElementalAnalysis row
    to an ExternalAnalysis record via external_analysis_id."""
    import io
    import openpyxl
    from backend.services.bulk_uploads.actlabs_titration_data import ActlabsRockTitrationService

    sample = SampleInfo(sample_id="TEST_ACTLABS_F1", rock_classification="Dunite")
    db_session.add(sample)
    db_session.flush()

    # Minimal ActLabs-format workbook:
    # row 0 — Report Number header
    # row 1 — Report Date header
    # row 2 — Analyte Symbol row (sample_id, Fe, SiO2)
    # row 3 — Unit Symbol row
    # row 4 — Detection Limit row
    # row 5 — Analysis Method row  ← data starts at row 6
    # row 6 — first data row
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Report Number", "12345"])
    ws.append(["Report Date", "2026-03-23"])
    ws.append(["sample_id", "Fe", "SiO2"])
    ws.append([None, "ppm", "%"])
    ws.append([None, "0.01", "0.01"])
    ws.append(["Analysis Method", "FUS-ICP", "FUS-ICP"])
    ws.append(["TEST_ACTLABS_F1", 45000.0, 38.5])
    buf = io.BytesIO()
    wb.save(buf)

    created, updated, skipped, errors = ActlabsRockTitrationService.import_excel(
        db_session, buf.getvalue()
    )
    assert errors == [], f"Unexpected errors: {errors}"
    assert created > 0

    rows = db_session.query(ElementalAnalysis).filter_by(sample_id="TEST_ACTLABS_F1").all()
    assert len(rows) > 0
    for row in rows:
        assert row.external_analysis_id is not None, (
            "external_analysis_id must be set on every ElementalAnalysis row"
        )
