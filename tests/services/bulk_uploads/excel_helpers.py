"""Minimal in-memory Excel fixture builders shared across bulk-upload service tests."""
from __future__ import annotations

import io
from typing import Any


def make_excel(headers: list[str], rows: list[list[Any]], sheet_name: str = "Sheet") -> bytes:
    """Return bytes of a minimal .xlsx file with one sheet."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_excel_multisheet(sheets: dict[str, tuple[list[str], list[list[Any]]]]) -> bytes:
    """
    Return bytes of a .xlsx file with named sheets.

    sheets = {"Dashboard": (["Col1", "Col2"], [[v1, v2], ...])}
    """
    import openpyxl

    wb = openpyxl.Workbook()
    first = True
    for sheet_name, (headers, rows) in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
