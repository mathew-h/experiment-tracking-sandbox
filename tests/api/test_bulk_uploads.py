"""Bulk-upload API endpoint tests.

Covers:
- Response shape for all 12 upload endpoints (mocked parsers)
- Auth rejection (401) for each endpoint family
- Required-file validation (422) for file-required endpoints
- Template downloads: 200 for valid types, 404 for no-template types
- Endpoint-specific: master-results sync (no file), timepoint-modifications form field
"""
from __future__ import annotations

import io
import sys
from unittest.mock import MagicMock, patch

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient

from backend.api.main import app
from backend.api.dependencies.db import get_db
from backend.auth.firebase_auth import verify_firebase_token, FirebaseUser


# ---------------------------------------------------------------------------
# Extra fixtures (unauth client for 401 tests)
# ---------------------------------------------------------------------------

# Re-export db_session and client from conftest implicitly via pytest fixture injection.


@pytest.fixture()
def unauth_client(db_session):
    """Client with DB override but verify_firebase_token raises 401."""

    def override_get_db():
        yield db_session

    async def no_auth():
        raise HTTPException(status_code=401, detail="Not authenticated")

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[verify_firebase_token] = no_auth
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_FAKE_XLSX = io.BytesIO(b"fake-excel-content")
_FILE_PARAM = {"file": ("test.xlsx", _FAKE_XLSX, "application/vnd.ms-excel")}


def _mock_upload_response(
    created=1, updated=0, skipped=0, errors=None, warnings=None, feedbacks=None
):
    return (created, updated, skipped, errors or [], feedbacks or [])


def _mock_upload_response_with_warnings(
    created=1, updated=0, skipped=0, errors=None, warnings=None, feedbacks=None
):
    return (created, updated, skipped, errors or [], warnings or [], {})


def _assert_upload_shape(body: dict):
    assert "created" in body
    assert "updated" in body
    assert "skipped" in body
    assert "errors" in body


# ---------------------------------------------------------------------------
# Existing endpoints — shape tests
# ---------------------------------------------------------------------------

def test_scalar_results_returns_upload_response_shape(client):
    mock_vc = MagicMock()
    mock_vc.SCALAR_RESULTS_TEMPLATE_HEADERS = []
    mock_svc = MagicMock()
    mock_svc.bulk_upsert_from_excel_ex.return_value = (1, 0, 0, [], [])

    fake_mod = MagicMock()
    fake_mod.ScalarResultsUploadService = mock_svc

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": mock_vc,
        "backend.services.bulk_uploads.scalar_results": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/scalar-results",
            files={"file": ("test.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_scalar_results_requires_file(client):
    resp = client.post("/api/bulk-uploads/scalar-results")
    assert resp.status_code == 422


def test_new_experiments_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.bulk_upsert_from_excel.return_value = (2, 0, 0, [], [], {})
    fake_mod = MagicMock()
    fake_mod.NewExperimentsUploadService = mock_svc

    with patch.dict(sys.modules, {"backend.services.bulk_uploads.new_experiments": fake_mod}):
        resp = client.post(
            "/api/bulk-uploads/new-experiments",
            files={"file": ("test.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_pxrf_returns_upload_response_shape(client):
    mock_vc = MagicMock()
    mock_pxrf = MagicMock()
    mock_pxrf.ingest_from_bytes.return_value = (3, 0, 0, [])
    fake_mod = MagicMock()
    fake_mod.PXRFUploadService = mock_pxrf

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": mock_vc,
        "backend.services.bulk_uploads.pxrf_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/pxrf",
            files={"file": ("test.csv", io.BytesIO(b"fake"), "text/csv")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_aeris_xrd_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.bulk_upsert_from_excel.return_value = (5, 2, 0, [])
    fake_mod = MagicMock()
    fake_mod.AerisXRDUploadService = mock_svc

    with patch.dict(sys.modules, {"backend.services.bulk_uploads.aeris_xrd": fake_mod}):
        resp = client.post(
            "/api/bulk-uploads/aeris-xrd",
            files={"file": ("test.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


# ---------------------------------------------------------------------------
# New endpoints — shape tests
# ---------------------------------------------------------------------------

def test_master_results_upload_returns_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.from_bytes.return_value = (3, 1, 0, [], [])
    fake_mod = MagicMock()
    fake_mod.MasterBulkUploadService = mock_svc

    with patch.dict(sys.modules, {"backend.services.bulk_uploads.master_bulk_upload": fake_mod}):
        resp = client.post(
            "/api/bulk-uploads/master-results",
            files={"file": ("master.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_master_results_sync_no_file_returns_response_shape(client):
    """POST to master-results without a file triggers sync_from_path."""
    mock_svc = MagicMock()
    mock_svc.sync_from_path.return_value = (2, 0, 0, [], [])
    fake_mod = MagicMock()
    fake_mod.MasterBulkUploadService = mock_svc

    with patch.dict(sys.modules, {"backend.services.bulk_uploads.master_bulk_upload": fake_mod}):
        resp = client.post("/api/bulk-uploads/master-results")
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_icp_oes_returns_upload_response_shape(client):
    stub_vc = MagicMock()
    stub_vc.ICP_FIXED_ELEMENT_FIELDS = ["fe", "si", "mg"]
    mock_icp = MagicMock()
    mock_icp.parse_and_process_icp_file.return_value = ([{"experiment_fk": 1}], [])
    mock_icp.bulk_create_icp_results.return_value = ([MagicMock()], [])
    fake_mod = MagicMock()
    fake_mod.ICPService = mock_icp

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": stub_vc,
        "backend.services.icp_service": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/icp-oes",
            files={"file": ("icp.csv", io.BytesIO(b"fake"), "text/csv")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_xrd_mineralogy_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.upload.return_value = (4, 1, 0, [])
    fake_mod = MagicMock()
    fake_mod.XRDAutoDetectService = mock_svc

    with patch.dict(sys.modules, {"backend.services.bulk_uploads.xrd_upload": fake_mod}):
        resp = client.post(
            "/api/bulk-uploads/xrd-mineralogy",
            files={"file": ("xrd.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_timepoint_modifications_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.bulk_set_from_bytes.return_value = (3, 0, [], [])
    fake_mod = MagicMock()
    fake_mod.TimepointModificationsService = mock_svc

    with patch.dict(sys.modules, {
        "backend.services.bulk_uploads.timepoint_modifications": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/timepoint-modifications",
            files={"file": ("mods.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
            data={"overwrite_existing": "false"},
        )
    assert resp.status_code == 200
    body = resp.json()
    _assert_upload_shape(body)
    assert body["created"] == 0  # timepoint endpoint always returns created=0


def test_rock_inventory_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.bulk_upsert_samples.return_value = (2, 1, 0, 0, [], [])
    fake_mod = MagicMock()
    fake_mod.RockInventoryService = mock_svc

    with patch.dict(sys.modules, {"backend.services.bulk_uploads.rock_inventory": fake_mod}):
        resp = client.post(
            "/api/bulk-uploads/rock-inventory",
            files={"file": ("rocks.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_chemical_inventory_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.bulk_upsert_from_excel.return_value = (5, 2, 0, [])
    fake_mod = MagicMock()
    fake_mod.ChemicalInventoryService = mock_svc

    with patch.dict(sys.modules, {
        "backend.services.bulk_uploads.chemical_inventory": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/chemical-inventory",
            files={"file": ("chems.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_elemental_composition_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.bulk_upsert_wide_from_excel.return_value = (10, 3, 0, [])
    fake_mod = MagicMock()
    fake_mod.ElementalCompositionService = mock_svc

    with patch.dict(sys.modules, {
        "backend.services.bulk_uploads.actlabs_titration_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/elemental-composition",
            files={"file": ("elem.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_actlabs_rock_returns_upload_response_shape(client):
    mock_svc = MagicMock()
    mock_svc.import_excel.return_value = (6, 1, 0, [])
    fake_mod = MagicMock()
    fake_mod.ActlabsRockTitrationService = mock_svc

    with patch.dict(sys.modules, {
        "backend.services.bulk_uploads.actlabs_titration_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/actlabs-rock",
            files={"file": ("rock.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


def test_experiment_status_returns_upload_response_shape(client):
    mock_preview = MagicMock()
    mock_preview.errors = []
    mock_preview.to_ongoing = []
    mock_preview.to_completed = []
    mock_preview.missing_ids = []

    mock_svc = MagicMock()
    mock_svc.preview_status_changes_from_excel.return_value = mock_preview
    mock_svc.apply_status_changes.return_value = (0, 0, {}, [])
    fake_mod = MagicMock()
    fake_mod.ExperimentStatusService = mock_svc

    with patch.dict(sys.modules, {
        "backend.services.bulk_uploads.experiment_status": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/experiment-status",
            files={"file": ("status.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
        )
    assert resp.status_code == 200
    _assert_upload_shape(resp.json())


# ---------------------------------------------------------------------------
# Auth rejection tests (sample of endpoints)
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("endpoint", [
    "/api/bulk-uploads/scalar-results",
    "/api/bulk-uploads/new-experiments",
    "/api/bulk-uploads/xrd-mineralogy",
    "/api/bulk-uploads/timepoint-modifications",
    "/api/bulk-uploads/rock-inventory",
    "/api/bulk-uploads/chemical-inventory",
    "/api/bulk-uploads/elemental-composition",
    "/api/bulk-uploads/actlabs-rock",
    "/api/bulk-uploads/experiment-status",
    "/api/bulk-uploads/icp-oes",
    "/api/bulk-uploads/aeris-xrd",
])
def test_upload_endpoint_requires_auth(unauth_client, endpoint):
    """Every upload endpoint rejects requests without a valid auth token."""
    resp = unauth_client.post(
        endpoint,
        files={"file": ("test.xlsx", io.BytesIO(b"fake"), "application/vnd.ms-excel")},
    )
    assert resp.status_code == 401


def test_master_results_sync_requires_auth(unauth_client):
    resp = unauth_client.post("/api/bulk-uploads/master-results")
    assert resp.status_code == 401


def test_template_download_requires_auth(unauth_client):
    resp = unauth_client.get("/api/bulk-uploads/templates/scalar-results")
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# Required file validation (422)
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("endpoint", [
    "/api/bulk-uploads/new-experiments",
    "/api/bulk-uploads/pxrf",
    "/api/bulk-uploads/aeris-xrd",
    "/api/bulk-uploads/xrd-mineralogy",
    "/api/bulk-uploads/rock-inventory",
    "/api/bulk-uploads/chemical-inventory",
    "/api/bulk-uploads/elemental-composition",
    "/api/bulk-uploads/actlabs-rock",
    "/api/bulk-uploads/experiment-status",
    "/api/bulk-uploads/icp-oes",
])
def test_upload_requires_file(client, endpoint):
    """Endpoints with required file return 422 when no file is supplied."""
    resp = client.post(endpoint)
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# Template download tests
# ---------------------------------------------------------------------------

_TEMPLATE_TYPES = [
    "scalar-results",
    "new-experiments",
    "rock-inventory",
    "chemical-inventory",
    "elemental-composition",
    "experiment-status",
    "timepoint-modifications",
    "xrd-mineralogy",
]

_NO_TEMPLATE_TYPES = [
    "master-results",
    "icp-oes",
    "aeris-xrd",
    "actlabs-rock",
    "pxrf",
]


@pytest.mark.parametrize("upload_type", _TEMPLATE_TYPES)
def test_template_download_returns_xlsx(client, upload_type):
    """Valid template types return 200 with an Excel content-type."""
    resp = client.get(f"/api/bulk-uploads/templates/{upload_type}")
    assert resp.status_code == 200, f"{upload_type}: {resp.text}"
    assert "spreadsheetml" in resp.headers.get("content-type", "")


@pytest.mark.parametrize("upload_type", _NO_TEMPLATE_TYPES)
def test_template_download_returns_404_for_no_template_types(client, upload_type):
    """Upload types without a template return 404."""
    resp = client.get(f"/api/bulk-uploads/templates/{upload_type}")
    assert resp.status_code == 404


def test_template_download_unknown_type_returns_404(client):
    resp = client.get("/api/bulk-uploads/templates/nonexistent-type")
    assert resp.status_code == 404


def test_xrd_template_experiment_mode(client):
    """XRD template accepts ?mode=experiment and returns an Excel file."""
    resp = client.get("/api/bulk-uploads/templates/xrd-mineralogy?mode=experiment")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")


# ---------------------------------------------------------------------------
# D1: AppConfig model upsert
# ---------------------------------------------------------------------------

def test_app_config_upsert(db_session):
    """AppConfig stores and retrieves key-value pairs."""
    from database.models.app_config import AppConfig

    cfg = AppConfig(key="test_key", value="test_value")
    db_session.add(cfg)
    db_session.flush()
    fetched = db_session.query(AppConfig).filter_by(key="test_key").first()
    assert fetched is not None
    assert fetched.value == "test_value"


def test_app_config_primary_key_is_key(db_session):
    """AppConfig uses key as PK — duplicate key raises integrity error."""
    from database.models.app_config import AppConfig
    import pytest

    db_session.add(AppConfig(key="dup_key", value="first"))
    db_session.flush()
    db_session.add(AppConfig(key="dup_key", value="second"))
    with pytest.raises(Exception):
        db_session.flush()


# ---------------------------------------------------------------------------
# D2: Master Results config endpoints
# ---------------------------------------------------------------------------

def test_get_master_results_config_returns_path(client):
    """GET /master-results/config returns JSON with a 'path' field."""
    r = client.get("/api/bulk-uploads/master-results/config")
    assert r.status_code == 200
    data = r.json()
    assert "path" in data


def test_patch_master_results_config_invalid_path(client):
    """PATCH /master-results/config rejects a nonexistent file path with 422."""
    r = client.patch(
        "/api/bulk-uploads/master-results/config",
        json={"path": "/nonexistent/path/to/file.xlsx"},
    )
    assert r.status_code == 422


def test_patch_master_results_config_valid_path(client, tmp_path):
    """PATCH /master-results/config accepts a real .xlsx file path and persists it."""
    import openpyxl

    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(p))

    r = client.patch(
        "/api/bulk-uploads/master-results/config",
        json={"path": str(p)},
    )
    assert r.status_code == 200
    assert r.json()["path"] == str(p)


def test_patch_master_results_config_persists_to_get(client, tmp_path):
    """After PATCH, GET returns the newly configured path."""
    import openpyxl

    p = tmp_path / "persist.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(p))

    client.patch("/api/bulk-uploads/master-results/config", json={"path": str(p)})

    r = client.get("/api/bulk-uploads/master-results/config")
    assert r.status_code == 200
    assert r.json()["path"] == str(p)


# ---------------------------------------------------------------------------
# pXRF reverse-match tests (Issue #28)
# ---------------------------------------------------------------------------

def _make_pxrf_excel_bytes(reading_nos: list) -> bytes:
    """Minimal pXRF Excel file with 'Reading No' column for reverse-match extraction."""
    import openpyxl  # noqa: PLC0415
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Reading No", "Fe", "Mg", "Ni", "Cu", "Si", "Co", "Mo", "Al", "Ca", "K", "Au"])
    for rno in reading_nos:
        ws.append([rno, 10.0, 1.0, 0.1, 0.1, 45.0, 0.1, 0.01, 8.0, 9.0, 1.0, 0.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_pxrf_upload_reevaluates_characterized_status(client, db_session):
    """After pXRF upload, a sample whose EA pxrf_reading_no matches the import becomes characterized."""
    from database import SampleInfo
    from database.models.analysis import ExternalAnalysis, PXRFReading

    # Sample is NOT characterized; has an EA pointing to reading_no "99"
    sample = SampleInfo(sample_id="REVTEST001", characterized=False)
    db_session.add(sample)
    db_session.flush()
    ea = ExternalAnalysis(
        sample_id="REVTEST001",
        analysis_type="pXRF",
        pxrf_reading_no="99",
    )
    db_session.add(ea)
    # PXRFReading "99" now exists (simulates what the upload would create)
    reading = PXRFReading(
        reading_no="99",
        fe=10.0, mg=1.0, ni=0.1, cu=0.1, si=45.0,
        co=0.1, mo=0.01, al=8.0, ca=9.0, k=1.0, au=0.0,
    )
    db_session.add(reading)
    db_session.flush()

    file_bytes = _make_pxrf_excel_bytes(["99"])
    fake_mod = MagicMock()
    fake_mod.PXRFUploadService.ingest_from_bytes.return_value = (1, 0, 0, [])

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": MagicMock(),
        "backend.services.bulk_uploads.pxrf_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/pxrf",
            files={"file": ("test.xlsx", io.BytesIO(file_bytes), "application/vnd.ms-excel")},
        )

    assert resp.status_code == 200
    db_session.refresh(sample)
    assert sample.characterized is True


def test_pxrf_upload_creates_modifications_log_entry(client, db_session):
    """ModificationsLog entry is created when characterized status changes after pXRF upload."""
    from database import SampleInfo
    from database.models.analysis import ExternalAnalysis, PXRFReading
    from database.models.experiments import ModificationsLog

    sample = SampleInfo(sample_id="REVTEST002", characterized=False)
    db_session.add(sample)
    db_session.flush()
    ea = ExternalAnalysis(
        sample_id="REVTEST002",
        analysis_type="pXRF",
        pxrf_reading_no="88",
    )
    db_session.add(ea)
    reading = PXRFReading(
        reading_no="88",
        fe=5.0, mg=2.0, ni=0.2, cu=0.2, si=40.0,
        co=0.2, mo=0.02, al=7.0, ca=8.0, k=2.0, au=0.0,
    )
    db_session.add(reading)
    db_session.flush()

    file_bytes = _make_pxrf_excel_bytes(["88"])
    fake_mod = MagicMock()
    fake_mod.PXRFUploadService.ingest_from_bytes.return_value = (1, 0, 0, [])

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": MagicMock(),
        "backend.services.bulk_uploads.pxrf_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/pxrf",
            files={"file": ("test.xlsx", io.BytesIO(file_bytes), "application/vnd.ms-excel")},
        )

    assert resp.status_code == 200
    log_entry = (
        db_session.query(ModificationsLog)
        .filter(ModificationsLog.sample_id == "REVTEST002")
        .first()
    )
    assert log_entry is not None
    assert log_entry.new_values.get("characterized") is True
    assert log_entry.old_values.get("characterized") is False


def test_pxrf_upload_message_includes_reevaluated_count(client, db_session):
    """Upload response message includes count of re-evaluated samples when > 0."""
    from database import SampleInfo
    from database.models.analysis import ExternalAnalysis, PXRFReading

    sample = SampleInfo(sample_id="REVTEST003", characterized=False)
    db_session.add(sample)
    db_session.flush()
    ea = ExternalAnalysis(
        sample_id="REVTEST003",
        analysis_type="pXRF",
        pxrf_reading_no="77",
    )
    db_session.add(ea)
    reading = PXRFReading(
        reading_no="77",
        fe=3.0, mg=3.0, ni=0.3, cu=0.3, si=35.0,
        co=0.3, mo=0.03, al=6.0, ca=7.0, k=3.0, au=0.0,
    )
    db_session.add(reading)
    db_session.flush()

    file_bytes = _make_pxrf_excel_bytes(["77"])
    fake_mod = MagicMock()
    fake_mod.PXRFUploadService.ingest_from_bytes.return_value = (1, 0, 0, [])

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": MagicMock(),
        "backend.services.bulk_uploads.pxrf_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/pxrf",
            files={"file": ("test.xlsx", io.BytesIO(file_bytes), "application/vnd.ms-excel")},
        )

    assert resp.status_code == 200
    body = resp.json()
    msg = body["message"].lower()
    assert "updated characterized" in msg or "1 sample" in msg


def test_pxrf_upload_no_change_when_no_matching_readings(client, db_session):
    """No re-evaluation count in message when uploaded readings don't match any sample EA."""
    from database import SampleInfo
    from database.models.analysis import ExternalAnalysis, PXRFReading

    # Sample has an EA for reading "55" but the upload contains reading "56" (no overlap)
    sample = SampleInfo(sample_id="REVTEST004")
    db_session.add(sample)
    db_session.flush()
    ea = ExternalAnalysis(
        sample_id="REVTEST004",
        analysis_type="pXRF",
        pxrf_reading_no="55",
    )
    db_session.add(ea)
    reading = PXRFReading(
        reading_no="56",
        fe=2.0, mg=4.0, ni=0.4, cu=0.4, si=30.0,
        co=0.4, mo=0.04, al=5.0, ca=6.0, k=4.0, au=0.0,
    )
    db_session.add(reading)
    db_session.flush()

    # Upload contains reading "56" which has no matching EA pxrf_reading_no
    file_bytes = _make_pxrf_excel_bytes(["56"])
    fake_mod = MagicMock()
    fake_mod.PXRFUploadService.ingest_from_bytes.return_value = (1, 0, 0, [])

    with patch.dict(sys.modules, {
        "frontend": MagicMock(),
        "frontend.config": MagicMock(),
        "frontend.config.variable_config": MagicMock(),
        "backend.services.bulk_uploads.pxrf_data": fake_mod,
    }):
        resp = client.post(
            "/api/bulk-uploads/pxrf",
            files={"file": ("test.xlsx", io.BytesIO(file_bytes), "application/vnd.ms-excel")},
        )

    assert resp.status_code == 200
    body = resp.json()
    # No samples matched the uploaded readings, so message has no "updated characterized"
    assert "updated characterized" not in body["message"].lower()
