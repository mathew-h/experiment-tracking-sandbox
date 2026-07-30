"""POST /api/bulk-uploads/experiment-deletion (issue #109, Phase 1).

Phase 1 is restricted to one hardcoded address. The single-user gate is the only
thing standing between an ordinary researcher and a hard, irreversible cascade
delete of an arbitrary list of experiments, so it is asserted from both sides:
the wrong user gets a 403 *and* the deletion service is never reached.

The cascade itself, per-row failure isolation and the audit row are covered by
tests/services/bulk_uploads/test_experiment_deletion_bulk.py; the service is
stubbed here so these tests exercise only the endpoint contract.
"""
from __future__ import annotations

import io
import sys
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient

from backend.api.dependencies.db import get_db
from backend.api.main import app
from backend.api.routers.bulk_uploads import BULK_DELETE_ALLOWED_EMAIL
from backend.auth.firebase_auth import FirebaseUser, verify_firebase_token

ENDPOINT = "/api/bulk-uploads/experiment-deletion"
_MODULE = "backend.services.bulk_uploads.experiment_deletion_bulk"


def _file(name: str = "delete.xlsx"):
    return {"file": (name, io.BytesIO(b"fake-excel"), "application/vnd.ms-excel")}


def _client_as(db_session, email: str) -> TestClient:
    def override_get_db():
        yield db_session

    def override_verify_token():
        return FirebaseUser(uid="uid", email=email, display_name="Test User")

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[verify_firebase_token] = override_verify_token
    return TestClient(app)


@pytest.fixture()
def owner_client(db_session):
    """Client authenticated as the one address allowed to bulk-delete."""
    with _client_as(db_session, BULK_DELETE_ALLOWED_EMAIL) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture()
def other_user_client(db_session):
    with _client_as(db_session, "someone.else@addisenergy.com") as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture()
def unauth_client(db_session):
    def override_get_db():
        yield db_session

    async def no_auth():
        raise HTTPException(status_code=401, detail="Not authenticated")

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[verify_firebase_token] = no_auth
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _stub_service(deleted=None, missing=None, failed=None, errors=None):
    """A stand-in for the deletion service module, with the real result shape."""
    result = MagicMock()
    result.deleted = deleted or []
    result.missing = missing or []
    result.failed = failed or []
    result.errors = errors or []

    fake_mod = MagicMock()
    fake_mod.delete_experiments_from_file.return_value = result
    return fake_mod


# ---------------------------------------------------------------------------
# Access control
# ---------------------------------------------------------------------------

def test_other_researcher_is_refused_and_the_service_is_never_reached(other_user_client):
    fake_mod = _stub_service(deleted=["SHOULD_NOT_HAPPEN"])

    with patch.dict(sys.modules, {_MODULE: fake_mod}):
        resp = other_user_client.post(ENDPOINT, files=_file())

    assert resp.status_code == 403
    fake_mod.delete_experiments_from_file.assert_not_called()


def test_the_gate_is_case_insensitive(db_session):
    """Firebase lower-cases addresses, but a mixed-case claim must not lock the
    owner out of their own cleanup tool."""
    fake_mod = _stub_service(deleted=["BDEL_API_001"])

    with _client_as(db_session, BULK_DELETE_ALLOWED_EMAIL.upper()) as client:
        with patch.dict(sys.modules, {_MODULE: fake_mod}):
            resp = client.post(ENDPOINT, files=_file())
    app.dependency_overrides.clear()

    assert resp.status_code == 200


def test_deletion_requires_authentication(unauth_client):
    resp = unauth_client.post(ENDPOINT, files=_file())
    assert resp.status_code == 401


def test_deletion_requires_a_file(owner_client):
    resp = owner_client.post(ENDPOINT)
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# Response contract
# ---------------------------------------------------------------------------

def test_deleted_missing_and_failed_all_reach_the_response(owner_client):
    fake_mod = _stub_service(
        deleted=["BDEL_API_001", "BDEL_API_002"],
        missing=["BDEL_API_TYPO"],
        failed=[{"experiment_id": "BDEL_API_LOCKED", "error": "row is locked"}],
    )

    with patch.dict(sys.modules, {_MODULE: fake_mod}):
        resp = owner_client.post(ENDPOINT, files=_file())

    assert resp.status_code == 200
    body = resp.json()
    assert body["updated"] == 2
    assert body["skipped"] == 1
    assert any("BDEL_API_LOCKED" in e and "row is locked" in e for e in body["errors"])

    feedback = body["feedbacks"][0]
    assert feedback["deleted"] == ["BDEL_API_001", "BDEL_API_002"]
    assert feedback["missing"] == ["BDEL_API_TYPO"]
    assert feedback["failed"] == [
        {"experiment_id": "BDEL_API_LOCKED", "error": "row is locked"}
    ]

    # The IDs have to be readable in the UI, which renders errors and warnings.
    assert any("BDEL_API_001" in w for w in body["warnings"])
    assert any("BDEL_API_TYPO" in w for w in body["warnings"])


def test_file_level_errors_report_zero_deletions(owner_client):
    fake_mod = _stub_service(errors=["Missing required column: 'experiment_id'"])

    with patch.dict(sys.modules, {_MODULE: fake_mod}):
        resp = owner_client.post(ENDPOINT, files=_file())

    assert resp.status_code == 200
    body = resp.json()
    assert body["updated"] == 0
    assert body["errors"] == ["Missing required column: 'experiment_id'"]


def test_the_filename_is_passed_through_so_csv_uploads_parse(owner_client):
    fake_mod = _stub_service(deleted=["BDEL_API_003"])

    with patch.dict(sys.modules, {_MODULE: fake_mod}):
        resp = owner_client.post(
            ENDPOINT,
            files={"file": (
                "cleanup.csv", io.BytesIO(b"experiment_id\nBDEL_API_003\n"), "text/csv",
            )},
        )

    assert resp.status_code == 200
    kwargs = fake_mod.delete_experiments_from_file.call_args.kwargs
    assert kwargs["filename"] == "cleanup.csv"
    assert kwargs["modified_by"] == BULK_DELETE_ALLOWED_EMAIL


def test_a_service_crash_returns_an_error_body_not_a_500(owner_client):
    fake_mod = MagicMock()
    fake_mod.delete_experiments_from_file.side_effect = RuntimeError("boom")

    with patch.dict(sys.modules, {_MODULE: fake_mod}):
        resp = owner_client.post(ENDPOINT, files=_file())

    assert resp.status_code == 200
    body = resp.json()
    assert body["updated"] == 0
    assert any("boom" in e for e in body["errors"])


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def test_template_is_a_single_column_experiment_id_sheet(owner_client):
    resp = owner_client.get("/api/bulk-uploads/templates/experiment-deletion")

    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb[wb.sheetnames[0]]
    assert [c.value for c in ws[1] if c.value is not None] == ["experiment_id"]
